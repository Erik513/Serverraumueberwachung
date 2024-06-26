import sqlite3
import csv
from datetime import datetime
#csv
from cryptography.fernet import Fernet
#pdf
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
#excel
import openpyxl
from openpyxl.utils import get_column_letter

#Datum
current_datetime = datetime.now()

class SensorDatabase: 
    def __init__(self, db_name="SensorMessungen.csv"):
        self.db_name = db_name
        self.create_table()

    def create_table(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS SensorMessungen (
                    ID INTEGER PRIMARY KEY,
                    Datum DATETIME,
                    Uhrzeit TEXT,
                    Luftfeuchtigkeit REAL,
                    Temperatur_in_C° REAL,
                    Abweichung TEXT,
                    Differenz TEXT
                )
            ''')
            cursor.execute('SELECT Datum FROM SensorMessungen WHERE ID = 1')
            first_date_in_db = cursor.fetchone()
            cursor.execute('SELECT Uhrzeit FROM SensorMessungen WHERE ID = 1')
            first_time_in_db = cursor.fetchone()
            if first_date_in_db and first_time_in_db:
                return f"Erstellungsdatum der Datenbank: {first_date_in_db[0]}, {first_time_in_db[0]}"
            else:
                return None
            
    def get_highest_hum(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT MAX(Luftfeuchtigkeit) FROM SensorMessungen')
            max_hum = cursor.fetchone()[0]
            return max_hum

    def get_lowest_hum(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT MIN(Luftfeuchtigkeit) FROM SensorMessungen')
            min_hum = cursor.fetchone()[0]
            return min_hum
        
    def get_highest_temp(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT MAX(Temperatur_in_C°) FROM SensorMessungen')
            max_temp = cursor.fetchone()[0]
            return max_temp
    
    def get_lowest_temp(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT MIN(Temperatur_in_C°) FROM SensorMessungen')
            min_temp = cursor.fetchone()[0]
            return min_temp

    def insert_measurement(self, r_temperature, r_humidity, temp_plus_minus, temperature_deviation):
        current_datetime = datetime.now()
        date_str = current_datetime.strftime('%d.%m.%Y')
        time_str = current_datetime.strftime('%X')

        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO SensorMessungen (Datum, Uhrzeit, Luftfeuchtigkeit, Temperatur_in_C°, Abweichung, Differenz)
                VALUES (? , ?, ?, ?, ?, ?)
            ''', (date_str, time_str, r_humidity, r_temperature, temp_plus_minus, temperature_deviation))
            print('Daten in die Datenbank gefüllt.')

    def export_to_csv(self, csv_filename="SensorMessungen.csv"):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM SensorMessungen')
            rows = cursor.fetchall()
            if rows:
                header = [description[0] for description in cursor.description]
                # Schreiben Sie die Daten in die CSV-Datei
                with open(csv_filename, 'w', newline='') as csv_file:
                    csv_writer = csv.writer(csv_file)
                    csv_writer.writerow(header)
                    csv_writer.writerows(rows)

                print(f'Datenbank in CSV-Datei exportiert: {csv_filename}')
            else:
                print('Die Datenbank ist leer. Keine CSV-Datei erstellt.')

    def get_all_measurements(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM SensorMessungen')
            return cursor.fetchall()
    
    @staticmethod
    def generate_password():
        import random
        import string
        characters = string.ascii_letters + string.digits
        password = ''.join(random.choice(characters) for _ in range(6))
        return password
    
    def create_pdf(self, password):        
        data = self.get_all_measurements()
        pdf_filename = "SensorMessungen.pdf"

        #Erstellen Sie das PDF-Dokument mit Verschlüsselung    
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter, encrypt=password)
        elements = []
        styles = getSampleStyleSheet()

        #Überschrift Text
        current_date = current_datetime.strftime('%d.%m.%Y (%X)')
        title_text = f"<u>Sensormessungen</u><br/><font size='8'>Stand: <font size='8' color='blue'>{current_date}</font></font>"
        title = Paragraph(title_text, styles['Title'])
        elements.append(title)

        elements.append(Spacer(1, 12))
        from main import temp_threshold_low, temp_threshold_high
        #Main Text
        additional_text = f"""
        {SensorDatabase.create_table(self)}<br/>
        <br/>
        Zulässige Temperaturen: {temp_threshold_low} - {temp_threshold_high} C° <br/>
        <br/>
        <u>Extremwerte (Niedrigste / Höchste):</u><br/>
        Luftfeuchtigkeit: {SensorDatabase.get_lowest_hum(self)} / {SensorDatabase.get_highest_hum(self)} % <br/>
        Temperatur: {SensorDatabase.get_lowest_temp(self)} / {SensorDatabase.get_highest_temp(self)} C° <br/>
        """
        main_text = Paragraph(additional_text, styles['Normal'])
        elements.append(main_text)

        #Leere Fläche
        elements.append(Spacer(1, doc.height / 20))

        #Eine Tabelle erstellen und Daten hinzufügen
        header = ["ID", "Datum", "Uhrzeit", "Luftfeuchtigkeit in %", "Temperatur in C°", "Abweichung", "Differenz"]
        table_data = [header] + [list(map(str, row)) for row in data]
        table = Table(table_data)

        #Gridlines zur Tabelle hinzufügen
        style = TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.black)])

        for row in range(1, len(table_data)):
            humidity_value = float(table_data[row][3])
            temperature_value = float(table_data[row][4])
            humidity_color_column = 3 
            temperature_color_column = 4 
            table_data[row][humidity_color_column] = self.get_humidity_color(humidity_value)
            table_data[row][temperature_color_column] = self.get_temperature_color(temperature_value)
            style.add('BACKGROUND', (humidity_color_column, row), (humidity_color_column, row), self.get_humidity_color(humidity_value))
            style.add('BACKGROUND', (temperature_color_column, row), (temperature_color_column, row), self.get_temperature_color(temperature_value))

        table.setStyle(style)

        # Fügen Sie die Tabelle zum PDF hinzu
        elements.append(table)

        # Erstellen Sie das PDF-Dokument
        doc.build(elements)

        print(f'Datenbank in verschlüsselter PDF-Datei exportiert: {pdf_filename}')
        return pdf_filename
    
    def get_humidity_color(self, humidity_value):
        if 0 <= humidity_value <= 10:
            return colors.red
        elif 10 < humidity_value <= 30:
            return colors.orange
        elif 30 < humidity_value <= 40:
            return colors.yellow
        elif 40 < humidity_value <= 60:
            return colors.green
        elif 60 < humidity_value <= 80:
            return colors.darkcyan
        elif 80 < humidity_value <= 90:
            return colors.blue
        elif 90 < humidity_value <= 100:
            return colors.darkblue
        else:
            return colors.white
    
    def get_temperature_color(self, temperature_value):
        from main import temp_threshold_low, temp_threshold_high
        if temperature_value < temp_threshold_low - 10:
            return colors.darkblue
        if temp_threshold_low - 10 <= temperature_value < temp_threshold_low - 5:
            return colors.blue
        if temp_threshold_low - 5 <= temperature_value < temp_threshold_low - 2:
            return colors.darkcyan
        elif temp_threshold_low - 2 <= temperature_value < temp_threshold_low:
            return colors.cyan
        elif temp_threshold_low <= temperature_value <= temp_threshold_high:
            return colors.green
        elif temp_threshold_high < temperature_value <= temp_threshold_high + 2:
            return colors.yellowgreen
        elif temp_threshold_high + 2 < temperature_value <= temp_threshold_high + 5:
            return colors.yellow
        elif temp_threshold_high + 5 < temperature_value <= temp_threshold_high + 10:
            return colors.orange
        elif temp_threshold_high + 10 < temperature_value:
            return colors.red
        else:
            return colors.white 
          
    def create_excel(self, excel_filename="SensorMessungen.xlsx"):
        data = self.get_all_measurements()
        
        # Erstellen Sie eine neue Arbeitsmappe
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Schreiben Sie die Überschriften
        header = ["ID", "Datum", "Uhrzeit", "Luftfeuchtigkeit in %", "Temperatur in C°", "Abweichung", "Differenz"]
        sheet.append(header)

        # Holen Sie alle Messungen aus der Datenbank
        data = self.get_all_measurements()

        # Schreiben Sie die Daten in die Tabelle
        for row in data:
            sheet.append(row)

        # Passen Sie die Spaltenbreite basierend auf dem Header an
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for cell in sheet[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Speichern Sie die Excel-Datei
        workbook.save(excel_filename)

        print(f'Datenbank in Excel-Datei exportiert: {excel_filename}')
        return excel_filename

    def delete_table(self):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute('DROP TABLE SensorMessungen')
            print(f'Datenbank gelöscht.')